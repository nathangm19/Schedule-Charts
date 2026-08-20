#!/usr/bin/env python3
"""
PLANNER EXTRACT — turns the takt flow schedule into a clean activity table.

Run this in Copilot 365 > Analyst with the Planner workbook attached, or
locally with:  python planner_extract.py

WHY A SCRIPT AND NOT A PROMPT
On the Planner an activity is a COLORED BAR, not a row of data. The fill
color IS the trade contractor - the word "Firstline" appears nowhere near
those bars. Three things have to be decoded, and all three are exact:

  1. Fill color -> trade.  About a third of the legend is stored as THEME
     colors (e.g. theme 5, tint 0), not hex. Those must be resolved
     through the workbook color scheme and an HLS lightness shift.
  2. Colored run -> one or more activities. A run of same-colored cells
     can hold two labels; the second label starts a second activity
     mid-run. Read the run as one bar and both get the same wrong dates.
  3. Column bands -> level and zone, carried down through merged cells.

OUTPUT
  Planner_Extract.csv / .xlsx, one row per activity bar:
    Level | Zone | Sub-Zone | Trade | Activity | Start | Finish |
    Workdays | Fill | Schedule Row | Basis

CONFIG - the only things that change between issues of the schedule.
"""

import glob, os, re, csv, colorsys, datetime, zipfile
import xml.etree.ElementTree as ET
import openpyxl

# --------------------------------------------------------------- CONFIG
SHEET       = None                  # None = auto-detect the takt sheet by its
                                    # rows-1/2 trade legend (on "Interiors
                                    # Planner" that is the "Lookahead" tab)
DATE_ROW    = 5                     # row holding the calendar dates
FIRST_ROW   = 13                    # first schedule body row
LEVEL_COL   = 2                     # B - Level band
ZONE_COL    = 3                     # C - Zone band
SUB_COL     = 4                     # D - Sub-zone band (rare)
LEGEND_ROWS = (1, 2)                # rows holding the trade legend
WEEK_START  = None                  # "2026-08-24" for a specific Monday;
                                    # None = the next Monday from today
# Outputs: Planner_Extract.xlsx/.csv (everything, + Calendar + Legend sheets)
#          Planner_Extract_Week.xlsx/.csv (the one week the WWP needs)
# -----------------------------------------------------------------------

def find_workbook():
    hits = [f for f in glob.glob("**/*.xls*", recursive=True)
            if "Extract" not in f and not os.path.basename(f).startswith("~")]
    if not hits:
        raise SystemExit("No .xlsx found. Attach the Planner workbook and re-run.")
    hits.sort(key=lambda f: -os.path.getsize(f))
    return hits[0]

# ---- theme colors live in the zip, not in openpyxl's object model ------
NS = "{http://schemas.openxmlformats.org/drawingml/2006/main}"
# xlsx theme indices are reordered vs the file: 0=lt1, 1=dk1, 2=lt2, 3=dk2
ORDER = ["lt1", "dk1", "lt2", "dk2", "accent1", "accent2", "accent3",
         "accent4", "accent5", "accent6", "hlink", "folHlink"]

def theme_palette(path):
    with zipfile.ZipFile(path) as z:
        name = next(n for n in z.namelist() if n.startswith("xl/theme/theme"))
        root = ET.fromstring(z.read(name))
    scheme = root.find(f".//{NS}clrScheme")
    out = {}
    for el in scheme:
        tag = el.tag.split("}")[1]
        srgb = el.find(f"{NS}srgbClr")
        sysc = el.find(f"{NS}sysClr")
        out[tag] = (srgb.get("val") if srgb is not None else sysc.get("lastClr")).upper()
    return {i: out[n] for i, n in enumerate(ORDER) if n in out}

def tint(hex6, t):
    """Excel tints in HLS: lighten toward white, darken toward black."""
    r, g, b = (int(hex6[i:i+2], 16) / 255 for i in (0, 2, 4))
    h, l, s = colorsys.rgb_to_hls(r, g, b)
    l = l if not t else (l * (1 + t) if t < 0 else l * (1 - t) + t)
    return "%02X%02X%02X" % tuple(round(c * 255) for c in colorsys.hls_to_rgb(h, l, s))

def fill_hex(cell, palette):
    f = cell.fill
    if f is None or f.fill_type in (None, "none"):
        return None
    fg = f.fgColor
    if fg.type == "rgb" and isinstance(fg.rgb, str):
        v = fg.rgb[-6:].upper()
        return None if v in ("FFFFFF", "000000") and cell.value is None else v
    if fg.type == "theme":
        base = palette.get(fg.theme)
        if base:
            try:    t = round(float(fg.tint or 0), 4)
            except Exception: t = 0.0
            return tint(base, t)
    return None

# ---- trade resolution --------------------------------------------------
# Colors that encode a STATUS rather than a contractor. Resolving these by
# nearest color produces confident nonsense - pale blue and pale green sit
# closer to each other than either sits to the right answer.
FLAG   = re.compile(r"crew availability|permit|fragnet", re.I)
DUR    = re.compile(r"\s*\(\s*\d+\s*d\s*;\s*\d+\s*c\s*\)\s*$", re.I)
ENERG  = re.compile(r"power energize|energize \d+v", re.I)
# Bright greens used as a completion highlight across many trades' work.
STATUS_FILLS = {"83E28E", "93E38D"}

def resolve_trade(hexc, legend, activity, row):
    if ENERG.search(activity):
        return "Veca", "Energization - those colors encode the power system (Normal / Critical / Life Safety / Optional / Equipment / UPS), not a trade"
    if hexc in STATUS_FILLS:
        return "Status (no trade color)", "Bar painted with a status color rather than a trade swatch - trade unknown, not guessed"
    if hexc == "FF0000":                      # legend has this twice
        if FLAG.search(activity):
            return "Issues", "Red flag - constraint, not a crew assignment"
        return "McKinstry", "Red is shared by McKinstry and Issues; work -> McKinstry"
    if hexc == "E2F0D9":                      # accent6 lighter 80%
        return "Wilkie", "Legacy Wilkie swatch, not in the current legend"
    if hexc in legend:
        return legend[hexc], "Legend match"
    return "UNRESOLVED", "Fill not in the legend - assign by hand"

def pick_sheet(wb, palette):
    """The takt sheet is the one whose rows 1-2 hold the trade legend:
    many distinct fill colors on labeled cells. Storefront EDD notes share
    one color, so they score 1 and can never win."""
    best, score = None, 0
    for ws in wb.worksheets:
        has_dates = any(isinstance(c.value, datetime.datetime) for c in ws[DATE_ROW])
        if not has_dates:
            continue
        fills = set()
        for r in LEGEND_ROWS:
            for c in ws[r]:
                if c.value not in (None, ""):
                    h = fill_hex(c, palette)
                    if h:
                        fills.add(h)
        if len(fills) > score:
            best, score = ws, len(fills)
    if best is None or score < 8:
        raise SystemExit("No sheet with a rows-1/2 trade legend found. "
                         "Is this the Interiors Planner workbook?")
    print(f'  takt sheet: "{best.title}" ({score} legend colors)')
    return best

def main():
    path = find_workbook()
    print(f"Reading {path}")
    palette = theme_palette(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[SHEET] if SHEET else pick_sheet(wb, palette)

    # 1. calendar -------------------------------------------------------
    dates = {c.column: c.value.date() for c in ws[DATE_ROW]
             if isinstance(c.value, datetime.datetime)}
    if not dates:
        raise SystemExit(f"No dates on row {DATE_ROW}.")
    dcols = sorted(dates)
    print(f"  {len(dcols)} workday columns, {dates[dcols[0]]} to {dates[dcols[-1]]}")

    # 2. legend ---------------------------------------------------------
    legend = {}
    for r in LEGEND_ROWS:
        for c in ws[r]:
            name = str(c.value).strip() if c.value else ""
            if not name:
                continue
            h = fill_hex(c, palette)
            if h and h not in legend:      # first swatch wins
                legend[h] = name
    print(f"  {len(legend)} legend colors")

    # 3. merged-cell map: a band's value only lives in its anchor -------
    anchor = {}
    for m in ws.merged_cells.ranges:
        for rr in range(m.min_row, m.max_row + 1):
            for cc in range(m.min_col, m.max_col + 1):
                anchor[(rr, cc)] = (m.min_row, m.min_col)

    def band(row, col):
        a = anchor.get((row, col), (row, col))
        # only accept the anchor if it sits in the SAME column, otherwise a
        # C:D merge makes the zone name masquerade as a sub-zone
        return ws.cell(*a).value if a[1] == col else None

    # 4. walk the body --------------------------------------------------
    level = zone = sub = ""
    rows = []
    for r in range(FIRST_ROW, ws.max_row + 1):
        lv, zn, sb = (band(r, LEVEL_COL), band(r, ZONE_COL), band(r, SUB_COL))
        if lv: level = re.sub(r"\s+", " ", str(lv)).strip()
        if zn: zone  = re.sub(r"\s+", " ", str(zn)).strip(); sub = ""
        if sb: sub   = re.sub(r"\s+", " ", str(sb)).strip()

        run = []                       # contiguous cells of one fill color
        def flush():
            if not run:
                return
            marks = [i for i, (_, _, v) in enumerate(run) if v]
            if not marks:              # colored but unlabeled - not an activity
                run.clear(); return
            for k, mi in enumerate(marks):
                s = 0 if k == 0 else mi                    # first label owns
                e = marks[k+1] - 1 if k+1 < len(marks) else len(run) - 1
                d0, d1 = dates[run[s][0]], dates[run[e][0]]
                act = DUR.sub("", re.sub(r"\s+", " ", str(run[mi][2])).strip()).strip()
                hexc = run[mi][1]
                trade, basis = resolve_trade(hexc, legend, act, r)
                # The Stairs and Energization blocks have a blank Level band on
                # this sheet, so a plain carry-down would file them under the
                # block above. Name them from their own content instead.
                lvl = level
                if ENERG.search(act):            lvl = "Energization"
                elif re.match(r"\s*Stair", zone): lvl = "Stairs"
                rows.append(dict(Level=lvl, Zone=zone, SubZone=sub, Trade=trade,
                                 Activity=act, Start=d0, Finish=d1,
                                 Workdays=(e - s + 1), Fill="#" + hexc,
                                 Row=r, Basis=basis))
            run.clear()

        for col in dcols:
            cell = ws.cell(r, col)
            h = fill_hex(cell, palette)
            v = str(cell.value).strip() if cell.value not in (None, "") else ""
            if h is None:
                flush(); continue
            if run and run[-1][1] != h:      # color change ends the bar
                flush()
            run.append((col, h, v))
        flush()

    # 5. the target week ------------------------------------------------
    if WEEK_START:
        wk_a = datetime.date.fromisoformat(WEEK_START)
    else:
        today = datetime.date.today()
        wk_a = today + datetime.timedelta(days=(7 - today.weekday()) % 7 or 7)
    wk_b = wk_a + datetime.timedelta(days=4)
    week = [r for r in rows if r["Start"] <= wk_b and r["Finish"] >= wk_a]
    print(f"  week of {wk_a} .. {wk_b}: {len(week)} activities")

    # 6. write ----------------------------------------------------------
    cols = ["Level","Zone","SubZone","Trade","Activity","Start","Finish",
            "Workdays","Fill","Row","Basis"]

    def write_csv(name, data):
        with open(name, "w", newline="", encoding="utf-8-sig") as fh:
            w = csv.DictWriter(fh, fieldnames=cols); w.writeheader(); w.writerows(data)

    def sheet_of(book, title, data):
        sh = book.active if book.active.title == "Sheet" else book.create_sheet()
        sh.title = title
        sh.append(cols)
        for r in data:
            sh.append([r[c] for c in cols])
        for i in range(1, len(cols) + 1):
            sh.cell(1, i).font = openpyxl.styles.Font(bold=True, color="FFFFFFFF")
            sh.cell(1, i).fill = openpyxl.styles.PatternFill("solid", fgColor="FF004F8C")
        for i in (6, 7):
            for rr in range(2, len(data) + 2):
                sh.cell(rr, i).number_format = "yyyy-mm-dd"
        for c, wdt in zip("ABCDEFGHIJK", [14, 40, 14, 22, 52, 12, 12, 10, 10, 8, 46]):
            sh.column_dimensions[c].width = wdt
        sh.freeze_panes = "A2"
        sh.auto_filter.ref = f"A1:K{len(data)+1}"
        return sh

    full = openpyxl.Workbook(); sheet_of(full, "Extract", rows)
    cal = full.create_sheet("Calendar"); cal.append(["Workday"])
    for c in dcols:
        cal.append([dates[c]]); cal.cell(cal.max_row, 1).number_format = "yyyy-mm-dd"
    lg = full.create_sheet("Legend"); lg.append(["Trade", "Fill"])
    for h, n in sorted(legend.items(), key=lambda kv: kv[1].lower()):
        lg.append([n, "#" + h])
    full.save("Planner_Extract.xlsx"); write_csv("Planner_Extract.csv", rows)

    wkb = openpyxl.Workbook(); sheet_of(wkb, "Extract", week)
    wkb.save("Planner_Extract_Week.xlsx"); write_csv("Planner_Extract_Week.csv", week)

    # 7. report ---------------------------------------------------------
    from collections import Counter
    print(f"\n{len(rows):,} activities -> Planner_Extract.xlsx  |  "
          f"{len(week):,} in the week -> Planner_Extract_Week.xlsx")
    bad = [r for r in rows if r["Trade"] == "UNRESOLVED"]
    stat = [r for r in rows if r["Trade"] == "Status (no trade color)"]
    if stat:
        print(f"\n{len(stat)} rows are status-colored - real work, trade not stated on the Planner.")
    if bad:
        print(f"\n{len(bad)} rows have a fill that is in neither the legend nor the status set.")
        print("These are the ONLY rows needing a human. Filter Trade = UNRESOLVED in the workbook:")
        for h, n in Counter(r["Fill"] for r in bad).most_common():
            ex = next(r["Activity"] for r in bad if r["Fill"] == h)
            print(f"   {h}  {n:>2} row(s)   e.g. {ex[:44]}")
    else:
        print("\nEvery bar resolved to a trade.")
    print("\nTop trades by crew-days:")
    cd = Counter()
    for r in rows: cd[r["Trade"]] += r["Workdays"]
    for t, n in cd.most_common(10):
        print(f"   {t:<26} {n:,}")

if __name__ == "__main__":
    main()
