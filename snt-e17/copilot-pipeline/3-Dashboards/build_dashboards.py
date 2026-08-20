#!/usr/bin/env python3
"""
DASHBOARD BUILDER - turns Planner_Extract.xlsx (step 1's FULL extract)
into the two self-contained HTML dashboards:

    SNT-Crew-Loading.html   crew loading by day, stacked by trade
    SNT-Gantt.html          waterfall Gantt, level -> zone, one row per
                            activity, filterable by trade

Run in Copilot 365 with Planner_Extract.xlsx AND the four template files
(crew_index.html, crew_app.js, gantt_head.html, gantt_app.js) attached,
or locally:  python build_dashboards.py

The templates ARE the locked design - this script only rebuilds the data
inside them. Colors come from the Planner's own trade legend (the Fill
column of the extract), never invented here.
"""
import datetime, glob, json, os, re, collections
import openpyxl

# --------------------------------------------------------------- CONFIG
EXTRACT = None        # None = auto-find Planner_Extract.xlsx (NOT the week file)
START   = None        # "2026-08-24" clips the window to that Monday;
                      # None = next Monday from today
# -----------------------------------------------------------------------

CANON={'Veca':'VECA','Pipefitter':'MMFS Mech Piping','Sheet Metal':'MMFS Sheet Metal',
 'Plumbing':'MMFS Plumbing','Medgas':'Mac-Miller Med Gas','MM-Controls':'Mac-Miller Controls',
 'Pro Clean':'ProClean','Kendell':'Kendall','Grazinni':'Grazzini','Acendent':'Ascendent',
 'NW Precast':'Northwest Precast','Fairweather':'Fairweather Masonry',
 'Flynn Roof':'Flynn','Flynn MP':'Flynn','Owner Furn.,':'Owner Furnished'}
LVLNORM={'Level E':'Level E','Level D':'Level D','Level C':'Level C','Level B':'Level B',
 'Lvl A':'Level A','LVL 1':'Level 1','LVL 3':'Level 3','LVL 6':'Level 6','LVL 7':'Level 7',
 'Level 8':'Level 8','LVL 9':'Level 9','LVL 10, 11, 12':'Level 10-12','LVL 14':'Level 14',
 'Level 14':'Level 14','LVL 15':'Level 15','Roof':'Roof','Elevators':'Elevators'}
LVLORDER=['Level E','Level D','Level C','Level B','Level A','Level 1','Level 3','Level 6',
 'Level 7','Level 8','Level 9','Level 10-12','Level 14','Level 15','Roof','Elevators',
 'Stairs','Energization']
# Consolidated change-order zone groups. First pattern that matches wins,
# so order matters: PR12 sits above EUTT to catch "EUTT/ PR12 Speakers Add".
GROUPS=[
 ('Eyewash Stations',      r'eyewash'),
 ('IDF Cooling Changes',   r'idf cooling'),
 ('Fluoroscopy',           r'fluoro'),
 ('Marion Vestibule',      r'margar'),
 ('PR12 Work',             r'pr\s?12'),
 ('EUTT',                  r'eutt'),
 ('Pharmacy',              r'pharmacy'),
 ('West Hoist Infill',     r'west hoist'),
 ('East Hoist Infill',     r'east hoist'),
 ('PR 11.5 Build Out',     r'pr\s?11\.5'),
 ('TEE Room Buildout',     r'tee room'),
 ('PR 14 Procedure Room',  r'pr\s?14'),
 ('Elevators',             r'elevator'),
]

def find_extract():
    hits=[f for f in glob.glob('**/*.xlsx', recursive=True)
          if 'extract' in os.path.basename(f).lower()
          and 'week' not in os.path.basename(f).lower()
          and not os.path.basename(f).startswith('~')]
    if not hits:
        raise SystemExit('Planner_Extract.xlsx not found. Attach step 1\'s FULL extract.')
    hits.sort(key=lambda f:-os.path.getsize(f))
    return hits[0]

path = EXTRACT or find_extract()
print('Reading', path)
wb = openpyxl.load_workbook(path, data_only=True)
if 'Extract' not in wb.sheetnames or 'Calendar' not in wb.sheetnames:
    raise SystemExit('%s is missing the Extract/Calendar sheets - run the '
                     'CURRENT planner_extract.py first (older versions did '
                     'not write the Calendar).' % path)

cal=[]
for row in wb['Calendar'].iter_rows(min_row=2, values_only=True):
    d=row[0]
    if isinstance(d, datetime.datetime): d=d.date()
    if isinstance(d, datetime.date): cal.append(d)
cal=sorted(set(cal)); pos={d:i for i,d in enumerate(cal)}
print(' %d workdays %s -> %s' % (len(cal), cal[0], cal[-1]))

hdr=[c.value for c in wb['Extract'][1]]
recs=[]; TCOL=collections.defaultdict(collections.Counter); skipped=0
for row in wb['Extract'].iter_rows(min_row=2, values_only=True):
    d=dict(zip(hdr,row))
    if not d.get('Activity'): continue
    s=d['Start'].date() if isinstance(d['Start'],datetime.datetime) else d['Start']
    e=d['Finish'].date() if isinstance(d['Finish'],datetime.datetime) else d['Finish']
    if s not in pos or e not in pos:
        skipped+=1; continue
    raw=str(d['Level']).strip()
    lvl='Stairs' if raw=='Stairs' else ('Energization' if raw=='Energization'
        else LVLNORM.get(raw, raw or 'Other'))
    zone='Energization' if lvl=='Energization' else (str(d['Zone']).strip() or 'Unzoned')
    if d.get('SubZone') and lvl!='Energization':
        zone='%s (%s)' % (zone, str(d['SubZone']).strip())
    trade=CANON.get(d['Trade'], d['Trade'])
    hexc=str(d.get('Fill') or '#999999').lstrip('#').upper()
    TCOL[trade][hexc]+=1
    recs.append((trade,lvl,zone,str(d['Activity']).strip(),pos[s],pos[e]))
if skipped: print(' NOTE: %d rows had dates outside the calendar - skipped' % skipped)
print(' %d activity bars' % len(recs))

trades=sorted({r[0] for r in recs})
tcol=['#'+TCOL[t].most_common(1)[0][0] for t in trades]
levels=[l for l in LVLORDER if l in {r[1] for r in recs}]
levels+= sorted({r[1] for r in recs}-set(levels))
zpairs=[]
for r in recs:
    p=(r[1],r[2])
    if p not in zpairs: zpairs.append(p)
zpairs.sort(key=lambda p:(levels.index(p[0]), p[1].lower()))
acts=[]; aidx={}
for r in recs:
    if r[3] not in aidx: aidx[r[3]]=len(acts); acts.append(r[3])
ti={t:i for i,t in enumerate(trades)}; zi={p:i for i,p in enumerate(zpairs)}
li={l:i for i,l in enumerate(levels)}
out={'days':[d.isoformat() for d in cal],'trades':trades,'tcol':tcol,'levels':levels,
     'zones':[[li[l],z] for l,z in zpairs],'acts':acts,
     'bars':[[ti[r[0]],zi[(r[1],r[2])],r[4],r[5],aidx[r[3]]] for r in recs]}
zcog=[]
for lv,z in out['zones']:
    g=-1
    for i,(nm,pat) in enumerate(GROUPS):
        if re.search(pat,z,re.I): g=i; break
    zcog.append(g)
out['cog']=[g[0] for g in GROUPS]; out['zcog']=zcog

# ------------------------------------------------------------- window
if START: mon=datetime.date.fromisoformat(START)
else:
    t=datetime.date.today()
    mon=t+datetime.timedelta(days=(7-t.weekday())%7 or 7)
lo=next((i for i,d in enumerate(cal) if d>=mon), 0)
tot=[0]*len(cal)
for b in out['bars']:
    for i in range(b[2],b[3]+1): tot[i]+=1
hi=max(i for i,v in enumerate(tot) if v>0)
out['clipped']=sum(1 for b in out['bars'] if b[2]<lo<=b[3])
out['days']=out['days'][lo:hi+1]
out['bars']=[[b[0],b[1],max(b[2]-lo,0),min(b[3]-lo,hi-lo),b[4]]
             for b in out['bars'] if not (b[3]<lo or b[2]>hi)]
span='%s to %s &middot; %d workdays' % (out['days'][0],out['days'][-1],len(out['days']))
data=json.dumps(out,separators=(',',':'))
print(' window %s | %d bars (%d clipped at start)' % (span.replace('&middot;','|'),
      len(out['bars']), out['clipped']))

# ------------------------------------------------------------- emit HTML
def need(name):
    if not os.path.exists(name):
        raise SystemExit('Template %s is missing - attach it and re-run.' % name)
    return open(name, encoding='utf-8').read()

wrote=[]
if os.path.exists('crew_index.html'):
    html=need('crew_index.html'); app=need('crew_app.js')
    html=re.sub(r'\d{4}-\d{2}-\d{2} to \d{4}-\d{2}-\d{2} &middot; \d+ workdays', span, html)
    open('SNT-Crew-Loading.html','w',encoding='utf-8').write(
        html+'<script>\nconst P='+data+';\n'+app+'\n</script>\n')
    wrote.append('SNT-Crew-Loading.html')
if os.path.exists('gantt_head.html'):
    html=need('gantt_head.html'); app=need('gantt_app.js')
    open('SNT-Gantt.html','w',encoding='utf-8').write(
        html+'\n<script>\nconst P='+data+';\n'+app+'\n</script>\n')
    wrote.append('SNT-Gantt.html')
if not wrote:
    raise SystemExit('No templates found (crew_index.html / gantt_head.html). '
                     'Attach the template files from this package.')
for w in wrote:
    print('WROTE %s  %s bytes' % (w, format(os.path.getsize(w),',')))
